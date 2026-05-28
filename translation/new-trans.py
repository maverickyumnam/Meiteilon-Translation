import os
import sys
import time
import pandas as pd
from sarvamai import SarvamAI
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# ================= CONFIG =================

MODEL = "sarvam-translate:v1"
SRC_LANG = "en-IN"
TGT_LANG = "mni-IN"

SRC_COL = "english"
OUT_COL = "manipuri"

ROW_DELAY = 0.05


# ================= HELPERS =================

def contains_bengali(text: str) -> bool:
    """Detect Bengali Unicode range."""
    return any(0x0980 <= ord(c) <= 0x09FF for c in text)


def create_client():
    """Ask user for API key and create Sarvam client."""
    api_key = input("\nEnter SARVAM API Key: ").strip()
    if not api_key:
        print("API key cannot be empty.")
        return create_client()
    return SarvamAI(api_subscription_key=api_key)


def translate_one(client, text):
    """Translate single row. If credit ends, ask new key and retry SAME row."""
    if not text.strip():
        return "", client

    while True:
        try:
            resp = client.text.translate(
                input=text,
                source_language_code=SRC_LANG,
                target_language_code=TGT_LANG,
                model=MODEL,
            )

            translated = (
                getattr(resp, "translated_text", None)
                or getattr(resp, "output_text", None)
            )

            if translated is None and isinstance(resp, dict):
                translated = (
                    resp.get("translated_text")
                    or resp.get("translation")
                    or resp.get("output")
                )

            return str(translated).strip(), client

        except Exception as e:
            error_text = str(e)

            # ===== CREDIT EXHAUSTED DETECTION =====
            if (
                "insufficient_quota_error" in error_text
                or "No credits available" in error_text
                or "status_code: 429" in error_text
            ):
                print("\n🚫 API credits exhausted!")
                print("🔑 Please enter a NEW API key to continue.")
                client = create_client()
                print("✅ New API key loaded. Resuming...\n")
                continue  # retry same row

            # Other temporary errors
            print(f"⚠ Temporary error: {e}")
            print("Retrying in 2 seconds...")
            time.sleep(2)


# ================= MAIN =================

def main():

    filename = input("Enter Excel filename (e.g. input.xlsx): ").strip()
    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"

    if not os.path.exists(filename):
        print("File not found.")
        sys.exit(1)

    out_file = "output_" + os.path.basename(filename)

    # ===== RESUME LOGIC =====
    if os.path.exists(out_file):
        print("\n🔄 Resuming from existing output file...")
        df = pd.read_excel(out_file, engine="openpyxl")
    else:
        df = pd.read_excel(filename, engine="openpyxl")

        if SRC_COL not in df.columns:
            first_col = df.columns[0]
            print(f"'{SRC_COL}' not found. Using '{first_col}' as English column.")
            df = df.rename(columns={first_col: SRC_COL})

        df[OUT_COL] = ""

    df[SRC_COL] = df[SRC_COL].fillna("").astype(str)
    df[OUT_COL] = df[OUT_COL].fillna("").astype(str)

    client = create_client()
    total = len(df)

    print("\n🚀 Starting translation...\n")

    for i in range(total):

        # Skip already translated rows
        if df.at[i, OUT_COL].strip():
            continue

        text = df.at[i, SRC_COL]

        print(f"Translating row {i+1}/{total}")

        translated, client = translate_one(client, text)
        df.at[i, OUT_COL] = translated

        # Save progress immediately (crash safe)
        df.to_excel(out_file, index=False, engine="openpyxl")

        time.sleep(ROW_DELAY)

    print("\n✅ Translation Complete.")

    # ===== Highlight Bengali Output =====
    wb = load_workbook(out_file)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    manipuri_col = headers.index(OUT_COL) + 1

    bengali_fill = PatternFill(
        start_color="FFF4CCCC",
        end_color="FFF4CCCC",
        fill_type="solid"
    )

    bengali_count = 0

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=manipuri_col)
        if cell.value and contains_bengali(str(cell.value)):
            cell.fill = bengali_fill
            bengali_count += 1

    wb.save(out_file)

    print("📁 Output file:", out_file)
    print("🔎 Bengali rows highlighted:", bengali_count)


if __name__ == "__main__":
    main()