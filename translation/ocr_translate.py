"""OCR a PDF into sentences and translate them with Azure Translator.

Usage examples:
  python translation/ocr_translate.py --pdf input.pdf --key YOUR_KEY --region eastus
  python translation/ocr_translate.py --pdf input.pdf --dry-run
"""
from pathlib import Path
import os
import sys

# Add project root directory to sys.path to enable direct execution
project_root = Path(__file__).resolve().parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

# Reconfigure stdout/stderr to avoid UnicodeEncodeError on Windows terminals
if sys.platform.startswith('win'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except Exception:
        pass

import time
import argparse

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from ocr.pdf_excel import process_pdf
from translation.azuretrans import (
    translate_one,
    get_azure_credentials,
    contains_bengali,
    AZURE_ENDPOINT,
    GLOBAL_ENDPOINT,
    SRC_COL,
    OUT_COL,
    ROW_DELAY,
)


def main():
    parser = argparse.ArgumentParser(description="OCR a PDF and translate extracted sentences")
    parser.add_argument("--pdf", "-p", help="PDF filename to OCR (from input/ if basename provided)")
    parser.add_argument("--input", "-i", help="Translate an existing XLSX from OCR outputs (lists files in output/ocr/ if omitted)")
    parser.add_argument("--out", "-o", help="Output XLSX filename (optional)")
    parser.add_argument("--endpoint", help="Azure endpoint URL")
    parser.add_argument("--key", help="Azure subscription key")
    parser.add_argument("--region", help="Azure resource region")
    parser.add_argument("--dry-run", action="store_true", help="Do not call Azure (test mode)")
    parser.add_argument("--delay", type=float, default=ROW_DELAY, help="Delay between rows (seconds)")
    args = parser.parse_args()

    INPUT_DIR = Path("input")
    OCR_OUTPUT_DIR = Path("output/ocr")
    TRANSLATION_OUTPUT_DIR = Path("output/translation")
    OCR_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    TRANSLATION_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    df = None
    CLEAN_OUTPUT_DIR = Path("output/clean")
    CLEAN_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if args.pdf:
        pdf_path = Path(args.pdf)
        if not pdf_path.exists():
            alt = INPUT_DIR / args.pdf
            if alt.exists():
                pdf_path = alt

        if not pdf_path.exists() or pdf_path.suffix.lower() != ".pdf":
            print("Invalid PDF file.")
            sys.exit(1)

        print(f"\n➡ OCRing PDF: {pdf_path}\n")
        sentences = process_pdf(pdf_path)

        if not sentences:
            print("No text extracted from PDF.")
            sys.exit(1)

        out_file = args.out or str(TRANSLATION_OUTPUT_DIR / f"{pdf_path.stem}_translated.xlsx")
        df = pd.DataFrame({SRC_COL: sentences, OUT_COL: [""] * len(sentences)})
    else:
        # translate existing xlsx from OCR outputs
        if args.input:
            input_path = Path(args.input)
            if not input_path.exists():
                alt = OCR_OUTPUT_DIR / args.input
                if alt.exists():
                    input_path = alt
        else:
            # List cleaned files (require cleaning before translation)
            candidates = sorted(CLEAN_OUTPUT_DIR.glob("*.xlsx"))
            if not candidates:
                print("No xlsx files found in 'output/clean/'. Run cleaner first or provide --pdf.")
                sys.exit(1)
            print("Available cleaned Excel files in output/clean:")
            for i, p in enumerate(candidates, start=1):
                print(f"  {i}. {p.name}")
            choice = input("Select number: ").strip()
            try:
                idx = int(choice) - 1
                input_path = candidates[idx]
            except Exception:
                print("Invalid selection")
                sys.exit(1)

        if not input_path.exists() or input_path.suffix.lower() != ".xlsx":
            print("Invalid input xlsx.")
            sys.exit(1)

        # if user chose from OCR outputs, try to find a cleaned version
        cleaned_candidate = CLEAN_OUTPUT_DIR / f"cleaned_{input_path.name}"
        if cleaned_candidate.exists():
            print(f"Found cleaned version: {cleaned_candidate.name}; using that as input.")
            input_path = cleaned_candidate

        out_file = args.out or str(TRANSLATION_OUTPUT_DIR / f"{input_path.stem}_translated.xlsx")
        df = pd.read_excel(str(input_path), engine="openpyxl")

    # ensure columns
    if SRC_COL not in df.columns:
        first_col = df.columns[0]
        df = df.rename(columns={first_col: SRC_COL})

    df[SRC_COL] = df[SRC_COL].fillna("").astype(str)
    if OUT_COL not in df.columns:
        df[OUT_COL] = ""
    df[OUT_COL] = df[OUT_COL].fillna("").astype(str)

    if args.dry_run:
        print("\n⚠ Running in dry-run mode: will NOT call Azure. Saving snapshot.")
        df.to_excel(out_file, index=False, engine="openpyxl")
        print("Saved:", out_file)
        return

    # Get credentials
    if args.key and args.region:
        endpoint = args.endpoint or AZURE_ENDPOINT
        key = args.key
        region = args.region
    else:
        endpoint, key, region = get_azure_credentials()

    total = len(df)
    print(f"\nTranslating {total} sentences via Azure...\n")

    for i in range(total):
        if str(df.at[i, OUT_COL]).strip():
            continue

        text = str(df.at[i, SRC_COL])
        print(f"Translating {i+1}/{total}")
        translated, (endpoint, key, region) = translate_one(endpoint, key, region, text)
        df.at[i, OUT_COL] = translated
        df.to_excel(out_file, index=False, engine="openpyxl")
        time.sleep(args.delay)

    print("\n✅ Translation complete. Performing Bengali-script highlighting...")

    wb = load_workbook(out_file)
    ws = wb.active

    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    bengali_fill = PatternFill(start_color="FFF4CCCC", end_color="FFF4CCCC", fill_type="solid")
    bengali_count = 0

    if OUT_COL in headers:
        manipuri_col = headers.index(OUT_COL) + 1
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=manipuri_col)
            if cell.value and contains_bengali(str(cell.value)):
                cell.fill = bengali_fill
                bengali_count += 1
        wb.save(out_file)
    else:
        wb.save(out_file)

    print("Output:", out_file)
    print("Bengali-script rows highlighted:", bengali_count)


if __name__ == "__main__":
    main()
