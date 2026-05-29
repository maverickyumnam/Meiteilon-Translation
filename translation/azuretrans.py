import os
import sys

# Reconfigure stdout/stderr to avoid UnicodeEncodeError on Windows terminals
if sys.platform.startswith('win'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except Exception:
        pass

import time
from typing import Tuple
import argparse
import socket
from urllib.parse import urlparse

import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ================= CONFIG =================

# Azure Translator API endpoint (default global)
AZURE_ENDPOINT = "https://centralindia.api.cognitive.microsofttranslator.com/"
GLOBAL_ENDPOINT = "https://api.cognitive.microsofttranslator.com/"
API_VERSION = "3.0"
SRC_LANG = "en"
TGT_LANG = "mni"          # Manipuri (Meitei Mayek) – change to "mni-Beng" if Bengali script is needed
SRC_COL = "english"
OUT_COL = "manipuri"
ROW_DELAY = 0.05

# ================= HELPERS =================

def contains_bengali(text: str) -> bool:
    """Detect Bengali Unicode range."""
    return any(0x0980 <= ord(c) <= 0x09FF for c in text)


def get_azure_credentials() -> Tuple[str, str, str]:
    """Ask user for Azure API key and region, return (endpoint, key, region)."""
    print("\nTo use Azure Translator, you need:")
    print("  - Subscription key (from your Translator resource)")
    print("  - Region (e.g., 'eastus', 'global' if no region is set)")

    key = input("Enter Azure Translator API key: ").strip()
    if not key:
        print("API key cannot be empty.")
        return get_azure_credentials()

    region = input("Enter resource region (e.g., eastus) [press Enter for 'global']: ").strip()
    if not region:
        region = "global"

    endpoint = input(f"Enter endpoint (press Enter for default '{AZURE_ENDPOINT}'): ").strip()
    if not endpoint:
        endpoint = AZURE_ENDPOINT
    # Ensure trailing slash
    if not endpoint.endswith("/"):
        endpoint += "/"

    return endpoint, key, region


def translate_one(endpoint: str, key: str, region: str, text: str) -> Tuple[str, Tuple[str, str, str]]:
    """Translate single row using Azure Translator. Return (translated_text, same credentials)."""
    if not text.strip():
        return "", (endpoint, key, region)

    url = f"{endpoint}translate?api-version={API_VERSION}&from={SRC_LANG}&to={TGT_LANG}"
    headers = {
        "Ocp-Apim-Subscription-Key": key,
        "Ocp-Apim-Subscription-Region": region,
        "Content-Type": "application/json"
    }
    body = [{"Text": text}]

    response = None
    while True:
        # ensure endpoint host resolves; if not, fall back to global endpoint
        try:
            parsed = urlparse(endpoint)
            host = parsed.netloc or parsed.path
            socket.getaddrinfo(host, 443)
        except Exception:
            if endpoint.rstrip('/') != GLOBAL_ENDPOINT.rstrip('/'):
                print(f"⚠ DNS failure for {endpoint!s}; falling back to global endpoint {GLOBAL_ENDPOINT}")
                endpoint = GLOBAL_ENDPOINT
                # update url for next request
                url = f"{endpoint}translate?api-version={API_VERSION}&from={SRC_LANG}&to={TGT_LANG}"
        try:
            response = requests.post(url, headers=headers, json=body, timeout=30)
            response.raise_for_status()  # raise HTTPError for bad status codes

            data = response.json()
            # response is list: [{'translations': [{'text': '...', 'to': 'mni'}]}]
            translated = data[0]["translations"][0]["text"]
            return translated.strip(), (endpoint, key, region)

        except requests.exceptions.HTTPError as e:
            if response is None:
                raise
            status_code = response.status_code
            error_text = response.text.lower()

            # ===== CREDIT / QUOTA EXHAUSTED DETECTION =====
            if status_code == 403 or status_code == 429 or "quota" in error_text:
                print("\n🚫 Azure credits exhausted or rate limit reached!")
                print("🔑 Please enter a NEW API key (with available quota) to continue.")
                endpoint, key, region = get_azure_credentials()
                print("✅ New Azure credentials loaded. Resuming...\n")
                continue   # retry same row

            # Other temporary errors (network, server error)
            print(f"⚠ Temporary error (HTTP {status_code}): {e}")
            print("Retrying in 2 seconds...")
            time.sleep(2)

        except Exception as e:
            print(f"⚠ Unexpected error: {e}")
            print("Retrying in 2 seconds...")
            time.sleep(2)


# ================= MAIN =================

def main():
    parser = argparse.ArgumentParser(description="Azure translation helper")
    parser.add_argument("--file", "-f", help="Excel filename (xlsx)")
    parser.add_argument("--endpoint", help="Azure endpoint URL")
    parser.add_argument("--key", help="Azure subscription key")
    parser.add_argument("--region", help="Azure resource region")
    parser.add_argument("--dry-run", action="store_true", help="Run without calling Azure (test mode)")
    args = parser.parse_args()

    if args.file:
        filename = args.file.strip()
    else:
        filename = input("Enter Excel filename (e.g. input.xlsx): ").strip()

    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"

    if not os.path.exists(filename):
        print("File not found.")
        sys.exit(1)

    base_dir = os.path.dirname(filename) or "."
    out_file = os.path.join(base_dir, "output_" + os.path.basename(filename))

    # ===== RESUME LOGIC =====
    if os.path.exists(out_file):
        print("\n🔄 Resuming from existing output file...")
        df = pd.read_excel(out_file, engine="openpyxl")
    else:
        df = pd.read_excel(filename, engine="openpyxl")

        if SRC_COL not in df.columns:
            first_col = df.columns[0]
            print(f"'{SRC_COL}' not found. Using '{first_col}' as English column.")
            df = df.rename(columns={first_col: SRC_COL})

        df[OUT_COL] = ""

    df[SRC_COL] = df[SRC_COL].fillna("").astype(str)
    df[OUT_COL] = df[OUT_COL].fillna("").astype(str)

    # Get Azure credentials (unless dry-run)
    if args.dry_run:
        print("\n⚠ Running in dry-run mode: will NOT call Azure.\n")
        total = len(df)
        # Save a snapshot so highlighting step has a workbook to open
        df.to_excel(out_file, index=False, engine="openpyxl")
    else:
        if args.key and args.region:
            endpoint = args.endpoint or AZURE_ENDPOINT
            key = args.key
            region = args.region
        else:
            endpoint, key, region = get_azure_credentials()
        total = len(df)

        print("\n🚀 Starting translation with Microsoft Azure Translator...\n")

        for i in range(total):
            # Skip already translated rows
            if str(df.at[i, OUT_COL]).strip():
                continue

            text = str(df.at[i, SRC_COL])

            print(f"Translating row {i+1}/{total}")

            translated, (endpoint, key, region) = translate_one(endpoint, key, region, text)
            df.at[i, OUT_COL] = translated

            # Save progress immediately (crash safe)
            df.to_excel(out_file, index=False, engine="openpyxl")

            time.sleep(ROW_DELAY)

    print("\n✅ Translation Complete.")

    # ===== Highlight Bengali Output =====
    wb = load_workbook(out_file)
    ws = wb.active

    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]

    bengali_fill = PatternFill(
        start_color="FFF4CCCC",
        end_color="FFF4CCCC",
        fill_type="solid"
    )

    bengali_count = 0

    if OUT_COL in headers:
        manipuri_col = headers.index(OUT_COL) + 1
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=manipuri_col)
            if cell.value and contains_bengali(str(cell.value)):
                cell.fill = bengali_fill
                bengali_count += 1

        wb.save(out_file)
    else:
        print(f"⚠ Column '{OUT_COL}' not found in output workbook; skipping Bengali highlighting.")
        wb.save(out_file)

    print("📁 Output file:", out_file)
    print("🔎 Bengali-script rows highlighted:", bengali_count)


if __name__ == "__main__":
    main()