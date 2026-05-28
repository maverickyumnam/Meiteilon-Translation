import os
import sys
import time
import pandas as pd
from sarvamai import SarvamAI

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# CONFIG 

MODEL = "sarvam-translate:v1"
SRC_LANG = "en-IN"
TGT_LANG = "mni-IN"

SRC_COL = "english"
OUT_COL = "manipuri"

MAX_RETRIES = 4
BASE_BACKOFF = 1.0   # seconds
ROW_DELAY = 0.05    # gentle on rate limits


# HELPERS 

def contains_bengali(text: str) -> bool:
    """Detect Bengali Unicode range."""
    return any(0x0980 <= ord(c) <= 0x09FF for c in text)


def translate_one(client: SarvamAI, text: str) -> str:
    """Translate one sentence with retries."""
    if not text or not text.strip():
        return ""

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = client.text.translate(
                input=text,
                source_language_code=SRC_LANG,
                target_language_code=TGT_LANG,
                model=MODEL,
            )

            # Extract translation safely
            translated = (
                getattr(resp, "translated_text", None)
                or getattr(resp, "output_text", None)
            )

            if translated is None and isinstance(resp, dict):
                translated = (
                    resp.get("translated_text")
                    or resp.get("translation")
                    or resp.get("output")
                )

            if translated is None:
                translated = str(resp)

            return str(translated).strip()

        except Exception as e:
            if attempt < MAX_RETRIES:
                wait = BASE_BACKOFF * (2 ** (attempt - 1))
                print(f"⚠ Error (attempt {attempt}/{MAX_RETRIES}): {e}")
                print(f"⏳ Retrying in {wait:.1f}s...")
                time.sleep(wait)
            else:
                print(f"❌ Failed after {MAX_RETRIES} attempts.")
                return ""


# MAIN 

def main():
    api_key = os.getenv("SARVAM_API_KEY")
    if not api_key:
        print("ERROR: SARVAM_API_KEY environment variable not set.")
        sys.exit(1)

    filename = input("Enter Excel filename (e.g. input.xlsx): ").strip()
    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"

    if not os.path.exists(filename):
        print("ERROR: File not found:", filename)
        sys.exit(1)

    df = pd.read_excel(filename, engine="openpyxl")

    # Ensure English column
    if SRC_COL not in df.columns:
        first_col = df.columns[0]
        print(f"'{SRC_COL}' not found. Using '{first_col}' as English.")
        df = df.rename(columns={first_col: SRC_COL})

    df[SRC_COL] = df[SRC_COL].fillna("").astype(str)

    client = SarvamAI(api_subscription_key=api_key)

    translations = []
    total = len(df)

    for i, text in enumerate(df[SRC_COL], start=1):
        print(f"Translating {i}/{total}")
        translations.append(translate_one(client, text))
        time.sleep(ROW_DELAY)

    df[OUT_COL] = translations

    out_file = "output_" + os.path.basename(filename)
    df[[SRC_COL, OUT_COL]].to_excel(out_file, index=False, engine="openpyxl")

    # Highlight Bengali Cells

    wb = load_workbook(out_file)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    manipuri_col = headers.index(OUT_COL) + 1

    bengali_fill = PatternFill(
        start_color="FFF4CCCC",
        end_color="FFF4CCCC",
        fill_type="solid"
    )

    bengali_count = 0

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=manipuri_col)
        if cell.value and contains_bengali(str(cell.value)):
            cell.fill = bengali_fill
            bengali_count += 1

    wb.save(out_file)

    print("\n DONE")
    print(" Output file:", out_file)
    print(f" Bengali rows highlighted: {bengali_count}")


if __name__ == "__main__":
    main()
